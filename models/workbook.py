import io
import requests
import logging
from copy import copy
from datetime import date
from functools import lru_cache
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.utils import get_column_letter
from core.order_parser import OrderParser
from core.template_loader import SHEET_FOAMING, SHEET_CARPENTER, SHEET_SALES

logger = logging.getLogger(__name__)

class WorkbookManager:
    """
    Base class for managing a monthly Excel workbook.
    Loads an existing workbook uploaded by the user and appends new sheets to it.
    A new workbook is created only when the month changes.

    template_bytes  — raw .xlsx bytes from TemplateLoader.raw_bytes
    template_sheet  — which sheet name inside the template to copy from
    existing_path   — path to the user-uploaded active workbook
    month_key       — e.g. "Jul/26", used for the output filename
    """

    # Excel's default column width/row height when no explicit dimension is set,
    # in the same units openpyxl stores them (chars / points). Needed because
    # column_dimensions / row_dimensions dicts are only populated for cells the
    # template explicitly sized — anything untouched falls back to these.
    _DEFAULT_COL_WIDTH_CHARS = 8.43
    _DEFAULT_ROW_HEIGHT_PT   = 15.0

    # A4 = paper size code 9 in the OOXML spec (openpyxl doesn't expose a
    # named constant for it). Orientation is set per subclass below.
    _PAPER_SIZE_A4     = 9
    _PAGE_ORIENTATION  = "landscape"   # overridden by Foaming (portrait)

    def __init__(self, existing_path: str, template_bytes: bytes,
                 template_sheet: str, month_key: str):
        self._source_path     = existing_path
        self._template_bytes  = template_bytes
        self._template_sheet  = template_sheet
        self._month_key       = month_key
        self._wb              = load_workbook(existing_path)

    @staticmethod
    @lru_cache(maxsize=None)
    def _load_template_workbook(template_bytes: bytes):
        """
        Parse the raw template bytes into an openpyxl Workbook once, cached
        on the bytes object itself.
        """
        return load_workbook(io.BytesIO(template_bytes))

    def _copy_template(self, sheet_name: str, max_row: int = None) -> Worksheet:
        """
        Load a fresh copy of the template workbook from bytes and manually
        replicate the relevant sheet into self._wb.
        openpyxl's copy_worksheet only works within a single workbook, so a
        temporary workbook can't be copy_worksheet'd directly into self._wb —
        instead we create a blank sheet here and transplant cell values,
        styles, merges, dimensions, and images one by one.

        max_row optionally caps how many of the template's rows get copied
        in.
        """
        tmp_wb = self._load_template_workbook(self._template_bytes)
        tmp_ws = tmp_wb[self._template_sheet]
        ws     = self._wb.create_sheet(title=sheet_name)

        self._copy_dimensions(tmp_ws, ws, max_row)
        self._copy_cells(tmp_ws, ws, max_row)
        self._copy_merged_cells(tmp_ws, ws)
        self._copy_images(tmp_ws, ws)

        ws.sheet_format = copy(tmp_ws.sheet_format)
        ws.freeze_panes  = tmp_ws.freeze_panes
        self._apply_print_setup(ws)
        return ws

    @staticmethod
    def _copy_dimensions(src: Worksheet, dst: Worksheet, max_row: int = None) -> None:
        for col_letter, dim in src.column_dimensions.items():
            dst.column_dimensions[col_letter].width  = dim.width
            dst.column_dimensions[col_letter].hidden = dim.hidden

        for row_idx, dim in src.row_dimensions.items():
            if max_row is not None and row_idx > max_row:
                continue
            dst_dim                = dst.row_dimensions[row_idx]
            dst_dim.height          = dim.height
            dst_dim.hidden          = dim.hidden
            dst_dim.outlineLevel    = dim.outlineLevel
            dst_dim.collapsed       = dim.collapsed

    @staticmethod
    def _copy_cells(src: Worksheet, dst: Worksheet, max_row: int = None) -> None:
        for row in src.iter_rows(max_row=max_row):
            for cell in row:
                new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font          = copy(cell.font)
                    new_cell.border        = copy(cell.border)
                    new_cell.fill          = copy(cell.fill)
                    new_cell.alignment     = copy(cell.alignment)
                    new_cell.protection    = copy(cell.protection)
                    new_cell.number_format = cell.number_format

    @staticmethod
    def _copy_merged_cells(src: Worksheet, dst: Worksheet) -> None:
        for merged_range in src.merged_cells.ranges:
            dst.merge_cells(str(merged_range))

    @staticmethod
    def _copy_images(src: Worksheet, dst: Worksheet) -> None:
        # e.g. the company logo on the foaming/carpenter/sales templates
        for img in src._images:
            dst.add_image(copy(img), img.anchor)

    @classmethod
    def _copy_row(cls, src_ws: Worksheet, src_row: int, dst_ws: Worksheet, dst_row: int,
                  max_col: int = None, copy_values: bool = True) -> None:
        """
        Copy one row's styles — and, unless copy_values=False, its values
        too — from src_row on src_ws to dst_row on dst_ws. src/dst can be
        the same worksheet or different ones, and the row index can shift.

        Generalizes two things this app needs: appending a plain new order
        row styled like the template's data row (copy_values=False — see
        _copy_row_style below), and transplanting a fully-formed row like
        the Carpenter Total row onto a different sheet entirely, values
        included (see CarpenterWorkbook._append_total_row).

        max_col defaults to src_ws's own column count rather than a
        caller-supplied constant, for the same reason _copy_row_style
        always has: Carpenter (9 cols) and Sales (12 cols) templates
        don't share a width, and a hardcoded value silently drifts out of
        sync if either template's column count ever changes.
        """
        if src_ws is dst_ws and src_row == dst_row:
            return

        if max_col is None:
            max_col = src_ws.max_column

        for col in range(1, max_col + 1):
            src_cell = src_ws.cell(row=src_row, column=col)
            dst_cell = dst_ws.cell(row=dst_row, column=col)
            if copy_values:
                dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font          = copy(src_cell.font)
                dst_cell.border        = copy(src_cell.border)
                dst_cell.fill          = copy(src_cell.fill)
                dst_cell.alignment     = copy(src_cell.alignment)
                dst_cell.protection    = copy(src_cell.protection)
                dst_cell.number_format = src_cell.number_format

        src_dim = src_ws.row_dimensions[src_row]
        dst_dim = dst_ws.row_dimensions[dst_row]
        dst_dim.height = src_dim.height

    @classmethod
    def _copy_row_style(cls, ws: Worksheet, src_row: int, dst_row: int, max_col: int = None) -> None:
        """
        Clone cell styles and row height from src_row onto dst_row, same
        worksheet, values left untouched. Used by Carpenter/Sales
        workbooks when appending a new order row beyond the template's
        first data row — ws.cell() creates plain, unstyled cells by
        default, so without this the appended rows would be missing the
        borders/fonts/height present on the template row.
        """
        cls._copy_row(ws, src_row, ws, dst_row, max_col, copy_values=False)

    def _next_empty_row(self, ws: Worksheet, check_column: int = 1) -> int:
        """
        First row from _DATA_START_ROW downward with nothing in
        check_column — i.e. the next free row to write an order into, or
        (reused by CarpenterWorkbook._append_total_row) the row directly
        below the last order already written.
        """
        row = self._DATA_START_ROW
        while ws.cell(row=row, column=check_column).value is not None:
            row += 1
        return row

    def has_sheet(self, sheet_name: str) -> bool:
        return sheet_name in self._wb.sheetnames

    def save(self) -> str:
        """
        Save the workbook to the same directory as the uploaded workbook.
        Returns the full path written.
        """
        directory = "/".join(self._source_path.replace("\\", "/").split("/")[:-1])
        path      = f"{directory}/{self._filename()}"
        self._wb.save(path)
        return path

    def _filename(self) -> str:
        raise NotImplementedError

    @staticmethod
    def date_to_sheet_name(d: date) -> str:
        """
        Format a date as dd.mm.yyyy for use as a carpenter/sales sheet name.
        e.g. date(2026, 6, 21) → "21.06.2026"
        """
        return f"{d.day:02d}.{d.month:02d}.{d.year}"

    @staticmethod
    def _col_width_to_px(width_chars: float) -> int:
        # Standard Excel char-width → pixel conversion for the default
        # (Calibri 11) font: pixels = round(chars * 7 + 5).
        return int(round(width_chars * 7 + 5))

    @staticmethod
    def _row_height_to_px(height_pt: float) -> int:
        # Points → pixels at 96 DPI (Excel's screen resolution assumption).
        return int(round(height_pt * 96 / 72))

    def _range_pixel_size(self, ws: Worksheet, min_col: int, max_col: int,
                       min_row: int, max_row: int) -> tuple[int, int]:
        """
        Total pixel width/height of a (possibly merged) cell range.

        Falls back to the sheet's own sheetFormatPr defaults
        (ws.sheet_format.defaultColWidth / defaultRowHeight) for any
        column/row that has no explicit dimension entry — these reflect
        the template's actual font/size and can differ from Excel's
        generic Calibri-11 defaults. The class constants below are only a last-resort
        fallback if the sheet doesn't specify its own defaults.

        Only used to estimate an aspect ratio now (see
        _letterbox_to_aspect_ratio) — the anchor itself is a TwoCellAnchor
        tied to real cell references, so any imprecision here just means
        slightly uneven padding, not a misplaced or overflowing image.
        """
        default_col_chars = ws.sheet_format.defaultColWidth or self._DEFAULT_COL_WIDTH_CHARS
        default_row_pts    = ws.sheet_format.defaultRowHeight or self._DEFAULT_ROW_HEIGHT_PT

        width_px = 0
        for col in range(min_col, max_col + 1):
            dim = ws.column_dimensions.get(get_column_letter(col))
            chars = dim.width if dim and dim.width else default_col_chars
            width_px += self._col_width_to_px(chars)

        height_px = 0
        for row in range(min_row, max_row + 1):
            dim = ws.row_dimensions.get(row)
            pts = dim.height if dim and dim.height else default_row_pts
            height_px += self._row_height_to_px(pts)

        return width_px, height_px

    def _embed_scaled_image(self, ws: Worksheet, image_url: str,
                             min_col: int, max_col: int,
                             min_row: int, max_row: int,
                             timeout: int = 10) -> None:
        """
        Download image_url and embed it (not linked — the actual bytes are
        stored in the workbook) into the given cell range, scaled to fit
        while preserving aspect ratio, centered, with leftover space left
        blank. Silently skips (with a logged warning) on any failure —
        a missing/broken product photo shouldn't block the whole order
        from being written to the workbook.

        The image is padded to the range's aspect ratio (see
        _letterbox_to_aspect_ratio) and anchored with a TwoCellAnchor
        spanning min_col:max_col / min_row:max_row.
        """
        if not image_url:
            logger.warning("No image URL provided — skipping image embed.")
            return

        try:
            resp = requests.get(image_url, timeout=timeout)
            resp.raise_for_status()
            src_img = PILImage.open(io.BytesIO(resp.content)).convert("RGBA")
        except Exception as exc:
            logger.warning(f"Could not fetch/read image from {image_url}: {exc}")
            return

        target_w, target_h = self._range_pixel_size(ws, min_col, max_col, min_row, max_row)

        target_w, target_h = self._range_pixel_size(ws, min_col, max_col, min_row, max_row)
        padded = self._letterbox_to_aspect_ratio(src_img, target_w / target_h)

        buffer = io.BytesIO()
        padded.save(buffer, format="PNG")
        buffer.seek(0)

        xl_img = XLImage(buffer)
        xl_img.anchor = TwoCellAnchor(
            editAs="twoCell",
            _from=AnchorMarker(col=min_col - 1, colOff=0, row=min_row - 1, rowOff=0),
            to=AnchorMarker(col=max_col, colOff=0, row=max_row, rowOff=0),
        )

        ws.add_image(xl_img)

    @staticmethod
    def _letterbox_to_aspect_ratio(img: PILImage.Image, target_ratio: float) -> PILImage.Image:
        """
        Pad img with transparent margin so its aspect ratio matches
        target_ratio, centering the original photo within the new
        canvas. This is what lets _embed_scaled_image hand Excel a plain
        stretch-to-fill TwoCellAnchor and still get scaled-to-fit,
        centered output — the padding does the centering/fitting work
        that used to be done with pixel offsets on the anchor itself.
        """
        src_w, src_h = img.size
        src_ratio = src_w / src_h

        if src_ratio > target_ratio:
            canvas_w, canvas_h = src_w, round(src_w / target_ratio)
        else:
            canvas_w, canvas_h = round(src_h * target_ratio), src_h

        canvas = PILImage.new("RGBA", (canvas_w, canvas_h), (0, 0, 0, 0))
        canvas.paste(img, ((canvas_w - src_w) // 2, (canvas_h - src_h) // 2), img)
        return canvas

    def _apply_print_setup(self, ws: Worksheet) -> None:
        """
        Every sheet this app generates is meant to be printed and handed
        to the shop floor, so force A4 + fit-to-one-page instead of
        Excel's default "actual size" behaviour, which can silently
        spill a sheet across several printed pages depending on the
        printer's default paper size/scale.
        """
        ws.page_setup.paperSize   = self._PAPER_SIZE_A4
        ws.page_setup.orientation = self._PAGE_ORIENTATION
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True


class FoamingWorkbook(WorkbookManager):
    """
    One sheet per work order, named by the order number suffix.
    e.g. G1/Jul/47 → sheet name "47"
    """

    # Merged image cell on the template — sofa photo goes here, scaled to fit.
    _IMAGE_MIN_COL, _IMAGE_MAX_COL = 2, 3     # B:C
    _IMAGE_MIN_ROW, _IMAGE_MAX_ROW = 26, 35   # rows 26–35

    _PAGE_ORIENTATION = "portrait"

    def __init__(self, existing_path: str, template_bytes: bytes, month_key: str):
        super().__init__(existing_path, template_bytes, SHEET_FOAMING, month_key)

    @staticmethod
    def last_order_number(existing_path: str) -> int | None:
        """
        Return the order number of the last sheet in an already-uploaded
        foaming workbook (sheets are named by order number — see add_order).
        Returns None if the last sheet isn't a plain number, e.g. a fresh
        workbook that has no order sheets yet.
        """
        wb = load_workbook(existing_path, read_only=True)
        last_sheet_name = wb.sheetnames[-1]
        try:
            return int(last_sheet_name)
        except ValueError:
            return None

    def add_order(self, wo_number: str, order_date: date, modified_delivery: date,
                  customer_name: str, order_id: str, product_name: str, qty: int, image_url: str = "") -> None:
        sheet_name = wo_number.split("/")[-1]   # "G1/Jul/47" → "47"

        if self.has_sheet(sheet_name):
            return   # order already exists in this workbook — skip

        ws = self._copy_template(sheet_name)
        self._fill(ws, wo_number, order_date, modified_delivery,
                   customer_name, order_id, product_name, qty)
        
        if image_url:
            self._embed_scaled_image(
                ws, image_url,
                self._IMAGE_MIN_COL, self._IMAGE_MAX_COL,
                self._IMAGE_MIN_ROW, self._IMAGE_MAX_ROW,
            )

    def _fill(self, ws: Worksheet, wo_number: str, order_date: date,
              modified_delivery: date, customer_name: str,
              order_id: str, product_name: str, qty: int) -> None:
        fmt = OrderParser.format_date

        ws["B4"]  = wo_number
        ws["E4"]  = fmt(order_date)
        ws["E5"]  = fmt(modified_delivery)
        ws["B8"]  = customer_name
        ws["C10"] = order_id
        ws["B13"] = product_name
        ws["E15"] = qty

    def _filename(self) -> str:
        month = self._month_key.replace("/", " ")
        return f"Test FO - {month}.xlsx"


class CarpenterWorkbook(WorkbookManager):
    """
    One sheet per order date, named dd.mm.yyyy.
    Multiple orders on the same date are appended as rows on the same sheet.

    The template's last row ("Total") is captured once
    (see _total_row_source) and transplanted onto each sheet this session
    creates, directly below that day's last order row, once save() runs
    (i.e. once all of that day's orders are in). Sheets already present in
    the uploaded workbook are left alone.
    """

    _DATA_START_ROW = 4

    def __init__(self, existing_path: str, template_bytes: bytes, month_key: str):
        super().__init__(existing_path, template_bytes, SHEET_CARPENTER, month_key)
        self._new_sheet_names: set[str] = set()   # sheets created THIS session

    def add_order(self, wo_number: str, modified_delivery: date, sku_id: str,
                  order_id: str, qty: int, order_date: date) -> None:
        sheet_name = self.date_to_sheet_name(order_date)

        if not self.has_sheet(sheet_name):
            _, total_row = self._total_row_source()
            self._copy_template(sheet_name, max_row=total_row - 1)
            self._new_sheet_names.add(sheet_name)

        ws       = self._wb[sheet_name]
        next_row = self._next_empty_row(ws)
        self._copy_row_style(ws, self._DATA_START_ROW, next_row)

        ws.cell(row=next_row, column=1).value = qty
        ws.cell(row=next_row, column=2).value = wo_number
        ws.cell(row=next_row, column=3).value = modified_delivery.strftime("%d/%m/%Y")
        ws.cell(row=next_row, column=4).value = sku_id
        # columns 5 (Fabric), 6 (Remark), 7 (Inches), 8 (Total Inches) — manual
        ws.cell(row=next_row, column=9).value = order_id

    def save(self) -> str:
        """
        Append the captured Total row to every sheet created THIS session
        — directly below its last order row, now that all of that sheet's
        orders are in — then hand off to the base save().
        """
        for sheet_name in self._new_sheet_names:
            self._append_total_row(self._wb[sheet_name])
        return super().save()
    
    def _append_total_row(self, ws: Worksheet) -> None:
        total_ws, total_row = self._total_row_source()
        self._copy_row(total_ws, total_row, ws, self._next_empty_row(ws))

    def _total_row_source(self) -> tuple[Worksheet, int]:
        """
        The template's own Carpenter worksheet plus the row index of its
        Total row (the template's last row).
        """
        tmp_wb = self._load_template_workbook(self._template_bytes)
        tmp_ws = tmp_wb[SHEET_CARPENTER]
        return tmp_ws, tmp_ws.max_row
    
    def _filename(self) -> str:
        month = self._month_key.replace("/", " ")
        return f"Test CA - {month}.xlsx"


class SalesWorkbook(WorkbookManager):
    """
    One sheet per order date, named dd.mm.yyyy.
    Multiple orders on the same date are appended as rows on the same sheet.
    """

    _DATA_START_ROW = 4
    _DATE_PLACEHOLDER = "[Order Confirmed Date]"

    def __init__(self, existing_path: str, template_bytes: bytes, month_key: str):
        super().__init__(existing_path, template_bytes, SHEET_SALES, month_key)

    def add_order(self, wo_number: str, modified_delivery: date,
                  customer_name: str, product_name: str, order_id: str,
                  qty: int, order_date: date) -> None:
        sheet_name = self.date_to_sheet_name(order_date)

        if not self.has_sheet(sheet_name):
            ws = self._copy_template(sheet_name)
            self._fill_header(ws, order_date)

        ws       = self._wb[sheet_name]
        next_row = self._next_empty_row(ws)
        self._copy_row_style(ws, self._DATA_START_ROW, next_row)

        ws.cell(row=next_row, column=1).value = next_row - self._DATA_START_ROW + 1   # serial number
        ws.cell(row=next_row, column=2).value = modified_delivery.strftime("%d/%m/%Y")
        ws.cell(row=next_row, column=3).value = wo_number
        ws.cell(row=next_row, column=4).value = qty
        ws.cell(row=next_row, column=5).value = customer_name
        ws.cell(row=next_row, column=6).value = product_name
        # columns 7 (Fabric), 8 (Remark) — manual
        ws.cell(row=next_row, column=9).value = order_id
        # columns 10 (Dispatch Date), 11 (Foaming Team), 12 (Carpenter Team) — manual

    def _fill_header(self, ws: Worksheet, order_date: date) -> None:
        """
        The template's date cell (A2) holds a literal placeholder —
        " Date : [Order Confirmed Date]" — swap it for the sheet's actual
        date the first time the sheet is created. All orders on this sheet
        share the same order_date (that's what date_to_sheet_name groups
        them by), so this only needs to run once per sheet, not per order.
        """
        cell = ws["A2"]
        if cell.value and self._DATE_PLACEHOLDER in cell.value:
            cell.value = cell.value.replace(
                self._DATE_PLACEHOLDER, OrderParser.format_date(order_date)
            )

    def _filename(self) -> str:
        month = self._month_key.replace("/", " ")
        return f"Test SO - {month}.xlsx"