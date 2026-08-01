import io
from openpyxl import load_workbook

from core.drive_cached_file import DriveCachedFile

SHEET_FOAMING   = "Foaming Template"
SHEET_CARPENTER = "Carpenter Template"
SHEET_SALES     = "Sales Summary"


class TemplateLoader(DriveCachedFile):
    """
    Downloads the template workbook from Google Drive and exposes
    the raw bytes so WorkbookManager can copy sheets from it.

    See DriveCachedFile for the caching behaviour — this class only
    supplies the file's identity and validates its sheets once loaded.
    """
    _FILE_ID      = "1UZak8H6roTIyS_t09wOoppET-4KtGXPS"
    _CHANGE_FIELD = "md5Checksum"
    _CACHE_NAME   = "template_cache"

    def _on_loaded(self) -> None:
        wb = load_workbook(io.BytesIO(self.raw_bytes))
        expected = {SHEET_FOAMING, SHEET_CARPENTER, SHEET_SALES}
        missing  = expected - set(wb.sheetnames)
        if missing:
            raise KeyError(
                f"Template workbook is missing sheets: {', '.join(sorted(missing))}\n"
                f"Expected: {', '.join(sorted(expected))}"
            )