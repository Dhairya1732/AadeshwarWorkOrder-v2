import io
from openpyxl import load_workbook

from core.drive_cached_file import DriveCachedFile

_SHEET_NAME = "Database Export Sheet"

_COL_PEPPERFRY_SKU_ID = "Pepperfry SKU ID"
_COL_CUSTOM_SKU = "Aadeshwar SKU ID"
_COL_FABRIC = "Fabric"

class SkuDatabase(DriveCachedFile):
    """
    Downloads the SKU/Fabric database sheet from Google Drive and exposes a
    lookup from Pepperfry "Your SKU ID" to (custom_sku, fabric).
    """
    _FILE_ID      = "11abJhgAGj7Y2io-by7SIuGKN_jUaY097uVMOIa3WTo8"
    _CHANGE_FIELD = "modifiedTime"
    _CACHE_NAME   = "sku_database_cache"

    _EXPORT_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(self):
        super().__init__()
        self._lookup: dict[str, tuple[str | None, str | None]] | None = None

    def get(self, sku_id: str) -> tuple[str | None, str | None]:
        if self._lookup is None:
            raise RuntimeError("SKU database not loaded — call fetch() first.")
        return self._lookup.get(sku_id.strip(), (None, None))

    def _download(self) -> bytes:
        return self._get(
            params={"mimeType": self._EXPORT_MIME, "key": self._API_KEY},
            purpose="download the SKU database",
            url=f"{self._drive_file_url}/export",
        ).content

    def _on_loaded(self) -> None:
        self._lookup = self._parse(self.raw_bytes)

    @staticmethod
    def _parse(raw_bytes: bytes) -> dict[str, tuple[str | None, str | None]]:
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        if _SHEET_NAME not in wb.sheetnames:
            raise KeyError(f"SKU database is missing expected sheet: {_SHEET_NAME}")
        ws = wb[_SHEET_NAME]

        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        col = {name: idx for idx, name in enumerate(header) if name}

        lookup: dict[str, tuple[str | None, str | None]] = {}
        for row in rows:
            sku_id = row[col[_COL_PEPPERFRY_SKU_ID]]
            if not sku_id:
                continue
            custom_sku = row[col[_COL_CUSTOM_SKU]]
            fabric     = row[col[_COL_FABRIC]]
            lookup[str(sku_id).strip()] = (
                str(custom_sku).strip() if custom_sku else None,
                str(fabric).strip() if fabric else None,
            )
        return lookup